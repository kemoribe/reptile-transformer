#!/usr/bin/env python3
"""Export all model rows from the approved Excel result tables to one CSV."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


OUTPUT_COLUMNS = [
    "source_file",
    "source_sheet",
    "dataset",
    "model",
    "R2",
    "RMSE",
    "MAE",
    "Pearson",
    "Spearman",
    "EF@1%",
    "EF@5%",
    "EF@10%",
    "ECE",
    "AUPR",
]

SHEET_DATASETS = {
    "ChEMBL数据集": "ChEMBL",
    "Davis数据集": "Davis",
    "BingdingBD数据集": "BindingDB",
    "KIBA数据集": "KIBA",
}

METRIC_HEADERS = {
    "R2": "R的平方",
    "RMSE": "RMSE",
    "MAE": "MAE",
    "Pearson": "Pearson",
    "Spearman": "Spearman",
    "EF@1%": "EF@1%",
    "EF@5%": "EF@5%",
    "EF@10%": "EF@10%",
    "ECE": "ECE",
    "AUPR": "AUPR",
}


def parse_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        match = re.search(r"[-+]?(?:\d+\.\d+|\d+)(?:[eE][-+]?\d+)?", value)
        if match:
            return float(match.group(0))
    return None


def export_tables(input_dir: Path, output_csv: Path) -> int:
    rows: list[dict[str, Any]] = []

    for workbook_path in sorted(input_dir.glob("*.xlsx")):
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for sheet_name, dataset in SHEET_DATASETS.items():
                if sheet_name not in workbook.sheetnames:
                    continue
                sheet = workbook[sheet_name]
                values = sheet.iter_rows(values_only=True)
                headers = next(values, None)
                if not headers:
                    continue

                header_indexes = {
                    str(header).strip(): index
                    for index, header in enumerate(headers)
                    if header is not None
                }
                model_index = header_indexes.get("模型")
                r2_index = header_indexes.get(METRIC_HEADERS["R2"])
                if model_index is None or r2_index is None:
                    continue

                for values_row in values:
                    model = values_row[model_index]
                    r2 = parse_number(values_row[r2_index])
                    if not model or r2 is None:
                        continue

                    row: dict[str, Any] = {
                        "source_file": workbook_path.name,
                        "source_sheet": sheet_name,
                        "dataset": dataset,
                        "model": str(model).strip(),
                        "R2": r2,
                    }
                    for output_name, source_header in METRIC_HEADERS.items():
                        if output_name == "R2":
                            continue
                        index = header_indexes.get(source_header)
                        row[output_name] = (
                            parse_number(values_row[index])
                            if index is not None and index < len(values_row)
                            else None
                        )
                    rows.append(row)
        finally:
            workbook.close()

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_dir", type=Path, help="Directory containing approved XLSX files")
    parser.add_argument("output_csv", type=Path, help="Combined CSV output path")
    args = parser.parse_args()

    row_count = export_tables(args.input_dir.resolve(), args.output_csv.resolve())
    print(f"Exported {row_count} rows to {args.output_csv.resolve()}")


if __name__ == "__main__":
    main()
